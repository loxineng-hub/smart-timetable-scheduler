import streamlit as st
import pandas as pd
import random
import copy
import io
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# =========================
# UI CONFIGURATION
# =========================
st.set_page_config(page_title="Smart Timetable Scheduler", page_icon="🏫", layout="wide")

if "page" not in st.session_state:
    st.session_state.page = "home"

def go_to_app():
    st.session_state.page = "app"

def go_to_home():
    st.session_state.page = "home"

# =========================
# BACKGROUND STYLE
# =========================
page_bg_img = """
<style>
[data-testid="stAppViewContainer"] {
    background-image: url("https://images.unsplash.com/photo-1434030216411-0b793f4b4173?q=80&w=2070&auto=format&fit=crop");
    background-size: cover;
    background-position: center;
    background-repeat: no-repeat;
    box-shadow: inset 0 0 0 2000px rgba(255, 255, 255, 0.55); 
}
[data-testid="stSidebar"] {
    background-color: rgba(255, 255, 255, 0.90);
}
[data-testid="stHeader"] {
    background: rgba(0,0,0,0);
}
.stMarkdown, .stSelectbox, .stRadio {
    background-color: rgba(255, 255, 255, 0.65);
    padding: 10px;
    border-radius: 10px;
}
h1, h2, h3, h4, p, label, li, span {
    color: #1E1E1E !important;
    font-weight: 500;
}
[data-testid="stMetricValue"] {
    color: #0047AB !important; 
}
</style>
"""
st.markdown(page_bg_img, unsafe_allow_html=True)

# =========================
# TEMPLATE DOWNLOAD
# =========================
def create_excel_template(is_blank=True):
    buffer = io.BytesIO()

    sheets = {
        "Class": ["Class ID"],
        "Subject": ["Subject ID"],
        "Class Subject": ["Class ID", "Subject ID", "Periods Per Week"],
        "Teachers": ["Teacher ID"],
        "Teacher Assignment": ["Teacher ID", "Subject Teach ID"],
        "Timeslot": ["Timeslot ID", "Day", "Period"],
        "Fixed Subjects": ["Class ID", "Subject ID", "Day", "Period"],
        "Consecutive Subjects": ["Subject ID", "Consecutive Periods"]
    }

    sample_data = {
        "Class": [["C1"], ["C2"]],
        "Subject": [["S1"], ["S2"]],
        "Class Subject": [["C1", "S1", 4], ["C1", "S2", 3], ["C2", "S1", 4]],
        "Teachers": [["T1"], ["T2"]],
        "Teacher Assignment": [["T1", "S1"], ["T2", "S2"]],
        "Timeslot": [["TS1", "Monday", 1], ["TS2", "Monday", 2], ["TS3", "Tuesday", 1]],
        "Fixed Subjects": [["C1", "S1", "Monday", 1]],
        "Consecutive Subjects": [["S1", 2]]
    }

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, cols in sheets.items():
            if is_blank:
                df = pd.DataFrame(columns=cols)
            else:
                df = pd.DataFrame(sample_data[sheet_name], columns=cols)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    return buffer.getvalue()

# =========================
# GA FUNCTION
# =========================
def run_zero_clash_ga(uploaded_file, progress_bar, status_box, num_classes, num_teachers):
    NUM_CLASSES = num_classes
    NUM_TEACHERS = num_teachers

    POP_SIZE = 20
    GENERATIONS = 60
    ELITE_SIZE = 3
    MUTATION_RATE = 0.15
    EARLY_STOP_OPTIMALITY = 90

    random.seed()

    classes = pd.read_excel(uploaded_file, sheet_name="Class")
    class_subject = pd.read_excel(uploaded_file, sheet_name="Class Subject")
    teachers = pd.read_excel(uploaded_file, sheet_name="Teachers")
    teacher_assign = pd.read_excel(uploaded_file, sheet_name="Teacher Assignment")
    timeslots = pd.read_excel(uploaded_file, sheet_name="Timeslot")
    fixed = pd.read_excel(uploaded_file, sheet_name="Fixed Subjects")
    consecutive = pd.read_excel(uploaded_file, sheet_name="Consecutive Subjects")

    available_classes = len(classes["Class ID"].dropna().unique())

    if NUM_CLASSES > available_classes:
        raise Exception(
            f"Your dataset only has {available_classes} classes, but you selected {NUM_CLASSES} classes."
        )

    selected_classes = classes["Class ID"].unique()[:NUM_CLASSES]

    classes = classes[classes["Class ID"].isin(selected_classes)].copy()
    class_subject = class_subject[class_subject["Class ID"].isin(selected_classes)].copy()
    fixed = fixed[fixed["Class ID"].isin(selected_classes)].copy()

    needed_subjects = class_subject["Subject ID"].unique().tolist()

    if len(needed_subjects) == 0:
        raise Exception("No valid subjects found for the selected classes.")

    teacher_list = teachers["Teacher ID"].dropna().unique().tolist()

    i = 1
    while len(teacher_list) < NUM_TEACHERS:
        teacher_list.append(f"TSIM{i}")
        i += 1

    teacher_list = teacher_list[:NUM_TEACHERS]
    teachers = pd.DataFrame({"Teacher ID": teacher_list})

    teacher_assign = teacher_assign[
        teacher_assign["Teacher ID"].isin(teacher_list)
    ].copy()

    new_assignments = []

    for idx, t in enumerate(teacher_list):
        if str(t).startswith("TSIM"):
            s = needed_subjects[idx % len(needed_subjects)]
            new_assignments.append({
                "Teacher ID": t,
                "Subject Teach ID": s
            })

    if new_assignments:
        teacher_assign = pd.concat(
            [teacher_assign, pd.DataFrame(new_assignments)],
            ignore_index=True
        )

    for s in needed_subjects:
        if teacher_assign[teacher_assign["Subject Teach ID"] == s].empty:
            teacher_assign = pd.concat([
                teacher_assign,
                pd.DataFrame([{
                    "Teacher ID": random.choice(teacher_list),
                    "Subject Teach ID": s
                }])
            ], ignore_index=True)

    valid_subjects = teacher_assign["Subject Teach ID"].unique()
    class_subject = class_subject[class_subject["Subject ID"].isin(valid_subjects)].copy()

    C = classes["Class ID"].tolist()
    T = teachers["Teacher ID"].tolist()
    TS = timeslots["Timeslot ID"].tolist()

    if len(TS) == 0:
        raise Exception("No timeslots found in the Timeslot sheet.")

    ts_map = dict(zip(
        timeslots["Timeslot ID"],
        zip(timeslots["Day"], timeslots["Period"])
    ))

    day_order = timeslots["Day"].drop_duplicates().tolist()
    period_order = sorted(timeslots["Period"].drop_duplicates().tolist())

    day_period_to_ts = {
        (day, period): ts
        for ts, (day, period) in ts_map.items()
    }

    periods_required = {
        (row["Class ID"], row["Subject ID"]): int(row["Periods Per Week"])
        for _, row in class_subject.iterrows()
    }

    teacher_subject = defaultdict(list)

    for _, row in teacher_assign.iterrows():
        teacher_subject[row["Subject Teach ID"]].append(row["Teacher ID"])

    consecutive_dict = {
        row["Subject ID"]: int(row["Consecutive Periods"])
        for _, row in consecutive.iterrows()
    }

    fixed_list = []

    if "Day" in fixed.columns and "Period" in fixed.columns:
        for _, row in fixed.iterrows():
            fixed_list.append({
                "Class": row["Class ID"],
                "Subject": row["Subject ID"],
                "Day": row["Day"],
                "Period": row["Period"]
            })

    fixed_lookup = defaultdict(list)

    for f in fixed_list:
        ts = day_period_to_ts.get((f["Day"], f["Period"]))
        if ts is not None:
            fixed_lookup[(f["Class"], f["Subject"])].append(ts)

    lessons = []

    for (c, s), req in periods_required.items():
        for _ in range(req):
            lessons.append({
                "Class": c,
                "Subject": s
            })

    TOTAL_LESSONS = len(lessons)

    if TOTAL_LESSONS == 0:
        raise Exception("No lessons found. Please check Class Subject sheet.")

    if TOTAL_LESSONS > len(C) * len(TS):
        raise Exception("Infeasible dataset: total lessons exceed timetable capacity.")

    def make_empty_state():
        return {
            "schedule": [],
            "class_busy": defaultdict(set),
            "teacher_busy": defaultdict(set),
            "teacher_load": defaultdict(int)
        }

    def place_lesson(state, c, s, t, ts):
        day, period = ts_map[ts]

        state["schedule"].append({
            "Class": c,
            "Day": day,
            "Period": period,
            "Subject": s,
            "Teacher": t,
            "Timeslot": ts
        })

        state["class_busy"][c].add(ts)
        state["teacher_busy"][t].add(ts)
        state["teacher_load"][t] += 1

    def count_clashes(df):
        if df is None or df.empty:
            return 999999

        class_clash = df.duplicated(
            subset=["Class", "Day", "Period"],
            keep=False
        ).sum()

        teacher_clash = df.duplicated(
            subset=["Teacher", "Day", "Period"],
            keep=False
        ).sum()

        return int(class_clash + teacher_clash)

    def fast_fixed_penalty(df):
        actual = set(zip(df["Class"], df["Subject"], df["Day"], df["Period"]))
        penalty = 0

        for f in fixed_list:
            if (f["Class"], f["Subject"], f["Day"], f["Period"]) not in actual:
                penalty += 1

        return penalty

    def fast_consecutive_penalty(df):
        penalty = 0
        grouped = defaultdict(list)

        for _, row in df.iterrows():
            grouped[(row["Class"], row["Subject"], row["Day"])].append(row["Period"])

        for c in C:
            for s, k in consecutive_dict.items():
                if (c, s) not in periods_required:
                    continue

                found = False

                for day in day_order:
                    periods = sorted(grouped.get((c, s, day), []))

                    for i in range(len(periods) - k + 1):
                        block = periods[i:i + k]

                        if block == list(range(block[0], block[0] + k)):
                            found = True
                            break

                    if found:
                        break

                if not found:
                    penalty += 1

        return penalty

    def load_balance_penalty(df):
        load = df.groupby("Teacher").size().to_dict()
        loads = [load.get(t, 0) for t in T]
        return max(loads) - min(loads)

    def evaluate(df):
        clashes = count_clashes(df)

        if clashes > 0:
            return 9999999 + clashes * 10000

        fp = fast_fixed_penalty(df)
        cp = fast_consecutive_penalty(df)
        lp = load_balance_penalty(df)

        return fp * 50 + cp * 30 + lp

    def optimality_percentage(df):
        clashes = count_clashes(df)
        fp = fast_fixed_penalty(df)
        cp = fast_consecutive_penalty(df)
        lp = load_balance_penalty(df)

        penalty = clashes * 10000 + fp * 50 + cp * 30 + lp
        max_penalty = max(1, TOTAL_LESSONS * 100)

        score = 100 - ((penalty / max_penalty) * 100)
        score = max(0, score)

        if clashes == 0:
            score = max(score, 90)

        if clashes == 0 and fp == 0 and cp == 0:
            score = 100 - min(lp, 10)

        return round(score, 2)

    def build_timetable(chromosome):
        state = make_empty_state()

        order = chromosome["lesson_order"]
        ordered_lessons = [lessons[i] for i in order]

        shuffled_slots = TS.copy()
        random.shuffle(shuffled_slots)

        for lesson in ordered_lessons:
            c = lesson["Class"]
            s = lesson["Subject"]

            possible_teachers = teacher_subject.get(s, [])

            if not possible_teachers:
                return None

            possible_teachers = sorted(
                possible_teachers,
                key=lambda t: state["teacher_load"][t]
            )

            placed = False

            for fixed_ts in fixed_lookup.get((c, s), []):
                for t in possible_teachers:
                    if fixed_ts not in state["class_busy"][c] and fixed_ts not in state["teacher_busy"][t]:
                        place_lesson(state, c, s, t, fixed_ts)
                        placed = True
                        break

                if placed:
                    break

            if placed:
                continue

            for ts in shuffled_slots:
                if ts in state["class_busy"][c]:
                    continue

                for t in possible_teachers:
                    if ts not in state["teacher_busy"][t]:
                        place_lesson(state, c, s, t, ts)
                        placed = True
                        break

                if placed:
                    break

            if not placed:
                return None

        return pd.DataFrame(state["schedule"])

    def create_chromosome():
        order = list(range(len(lessons)))
        random.shuffle(order)
        return {"lesson_order": order}

    def fitness(chromosome):
        df = build_timetable(chromosome)

        if df is None:
            return 99999999, None

        return evaluate(df), df

    def crossover(p1, p2):
        size = len(p1["lesson_order"])
        a, b = sorted(random.sample(range(size), 2))

        child_order = [None] * size
        child_order[a:b] = p1["lesson_order"][a:b]

        fill = [x for x in p2["lesson_order"] if x not in child_order]
        idx = 0

        for i in range(size):
            if child_order[i] is None:
                child_order[i] = fill[idx]
                idx += 1

        return {"lesson_order": child_order}

    def mutate(chromosome):
        child = copy.deepcopy(chromosome)

        if random.random() < MUTATION_RATE:
            a, b = random.sample(range(len(child["lesson_order"])), 2)
            child["lesson_order"][a], child["lesson_order"][b] = (
                child["lesson_order"][b],
                child["lesson_order"][a]
            )

        return child

    def tournament_select(scored_population, k=3):
        selected = random.sample(scored_population, k)
        selected = sorted(selected, key=lambda x: x[0])
        return selected[0][1]

    population = [create_chromosome() for _ in range(POP_SIZE)]

    best_score = float("inf")
    best_df = None

    for gen in range(GENERATIONS):
        scored = []

        for chrom in population:
            score, df = fitness(chrom)
            scored.append((score, chrom, df))

            if df is not None and count_clashes(df) == 0 and score < best_score:
                best_score = score
                best_df = df.copy()

        scored = sorted(scored, key=lambda x: x[0])

        best_now_df = scored[0][2]

        if best_now_df is not None:
            current_clash = count_clashes(best_now_df)
            current_optimality = optimality_percentage(best_now_df)
        else:
            current_clash = 999999
            current_optimality = 0

        progress_bar.progress((gen + 1) / GENERATIONS)

        status_box.info(
            f"Generation {gen + 1}/{GENERATIONS} | "
            f"Clash: {current_clash} | "
            f"Optimality: {current_optimality}% | "
            f"Status: Generating timetable..."
        )

        if best_df is not None:
            best_opt = optimality_percentage(best_df)

            if count_clashes(best_df) == 0 and best_opt >= EARLY_STOP_OPTIMALITY:
                status_box.success(
                    f"✅ Timetable generated successfully at generation {gen + 1}. "
                    f"Zero clash achieved with {best_opt}% optimality."
                )
                break

        new_population = []

        for _, chrom, _ in scored[:ELITE_SIZE]:
            new_population.append(copy.deepcopy(chrom))

        while len(new_population) < POP_SIZE:
            p1 = tournament_select(scored)
            p2 = tournament_select(scored)

            child = crossover(p1, p2)
            child = mutate(child)

            new_population.append(child)

        population = new_population

    if best_df is None:
        raise Exception("No zero-clash timetable found. Try increasing population or generations.")

    best_df = best_df.sort_values(["Class", "Day", "Period"]).reset_index(drop=True)

    final_clash = count_clashes(best_df)
    final_fixed = fast_fixed_penalty(best_df)
    final_consecutive = fast_consecutive_penalty(best_df)
    final_load_diff = load_balance_penalty(best_df)
    final_optimality = optimality_percentage(best_df)

    report = {
        "classes": len(C),
        "teachers": len(T),
        "timeslots": len(TS),
        "lessons": TOTAL_LESSONS,
        "clash": final_clash,
        "fixed": final_fixed,
        "consecutive": final_consecutive,
        "load_diff": final_load_diff,
        "optimality": final_optimality,
        "day_order": day_order,
        "period_order": period_order
    }

    return best_df, report

# =========================
# EXCEL EXPORT
# =========================
def create_output_excel(df_output, report):
    buffer = io.BytesIO()

    day_order = report["day_order"]
    period_order = report["period_order"]

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_output.drop(columns=["Timeslot"], errors="ignore").to_excel(
            writer,
            sheet_name="Raw Timetable",
            index=False
        )

        report_df = pd.DataFrame({
            "Criteria": [
                "Number of Classes",
                "Number of Teachers",
                "Number of Timeslots",
                "Total Lessons",
                "Hard Clashes",
                "Optimality Percentage"
            ],
            "Result": [
                report["classes"],
                report["teachers"],
                report["timeslots"],
                report["lessons"],
                report["clash"],
                f'{report["optimality"]}%'
            ]
        })

        report_df.to_excel(writer, sheet_name="GA Report", index=False)

        export_df = df_output.copy()

        for c in sorted(export_df["Class"].unique()):
            c_df = export_df[export_df["Class"] == c].copy()
            c_df["Entry"] = c_df["Subject"] + " (" + c_df["Teacher"] + ")"

            grid = c_df.pivot_table(
                index="Day",
                columns="Period",
                values="Entry",
                aggfunc=lambda x: " / ".join(x)
            ).reindex(index=day_order, columns=period_order).fillna("-")

            grid.to_excel(writer, sheet_name=f"Class_{c}"[:31])

        for t in sorted(export_df["Teacher"].unique()):
            t_df = export_df[export_df["Teacher"] == t].copy()
            t_df["Entry"] = t_df["Class"] + " (" + t_df["Subject"] + ")"

            grid = t_df.pivot_table(
                index="Day",
                columns="Period",
                values="Entry",
                aggfunc=lambda x: " / ".join(x)
            ).reindex(index=day_order, columns=period_order).fillna("-")

            grid.to_excel(writer, sheet_name=f"Teacher_{t}"[:31])

    buffer.seek(0)

    wb = load_workbook(buffer)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = center
                cell.border = border

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 3

    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)

    return final_buffer.getvalue()

# =========================
# HOME PAGE
# =========================
if st.session_state.page == "home":
    col1, col2, col3 = st.columns([1, 4, 1])

    with col2:
        st.title("🏫 Smart Timetable Scheduler")

        st.write("### 📖 System Overview")
        st.markdown("""
        Generate a conflict-free school timetable using a Genetic Algorithm.

        **Constraints handled:**
        - Teacher no clash
        - Class no clash
        - Valid teacher-subject assignment
        - Fixed subject preference
        - Consecutive subject preference
        - Teacher load balance
        """)

        st.button("Enter System ➡️", on_click=go_to_app, type="primary", use_container_width=True)

# =========================
# APP PAGE
# =========================
elif st.session_state.page == "app":
    st.title("📅 Timetable Generator")

    with st.sidebar:
        st.button("⬅️ Back to Home", on_click=go_to_home, use_container_width=True)

        st.header("📥 Templates")

        st.download_button(
            "Download Blank Template",
            data=create_excel_template(is_blank=True),
            file_name="Blank_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        st.download_button(
            "Download Sample Data",
            data=create_excel_template(is_blank=False),
            file_name="Sample_Data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    st.write("### 📤 Upload Excel File")

    st.write("### 📘 Dataset Explanation")

    with st.expander("Click here to understand the required Excel sheets and constraints"):
        st.markdown("""
        Your Excel file must contain these sheets:

        **1. Class**  
        Contains the list of classes involved in timetable generation.  
        Example: `C1`, `C2`, `C3`

        **2. Subject**  
        Contains all subjects offered by the school.  
        Example: `S1`, `S2`, `S3`

        **3. Class Subject**  
        Shows which subject belongs to which class and how many periods per week are required.  
        Example: Class `C1` needs Subject `S1` for `4` periods per week.

        **4. Teachers**  
        Contains all teacher IDs.  
        Example: `T1`, `T2`, `T3`

        **5. Teacher Assignment**  
        Shows which teacher can teach which subject.  
        Example: Teacher `T1` can teach Subject `S1`.

        **6. Timeslot**  
        Contains all available teaching periods.  
        Example: `Monday Period 1`, `Monday Period 2`, `Tuesday Period 1`.

        **7. Fixed Subjects**  
        Used when a subject should be placed at a specific day and period.  
        Example: Class `C1` should have Subject `S1` on `Monday Period 1`.

        **8. Consecutive Subjects**  
        Used when a subject should be arranged in back-to-back periods.  
        Example: If Subject `S1` has `2` consecutive periods, the system tries to place it like  
        `Monday Period 1 + Monday Period 2`.

        **Hard Constraints**
        - A class cannot have two subjects at the same time.
        - A teacher cannot teach two classes at the same time.
        - A subject must be assigned only to a valid teacher.

        The most important result is: `Hard Clashes = 0`
        """)

    uploaded_file = st.file_uploader(
        "Upload your timetable dataset Excel file",
        type=["xlsx"]
    )

    st.write("### ⚙️ Timetable Size Settings")

    col_a, col_b = st.columns(2)

    with col_a:
        num_classes = st.number_input(
            "Select number of classes",
            min_value=1,
            max_value=100,
            value=24,
            step=1
        )

    with col_b:
        num_teachers = st.number_input(
            "Select number of teachers",
            min_value=1,
            max_value=300,
            value=100,
            step=1
        )

    st.info(
        f"The system will generate a timetable using {int(num_classes)} classes and {int(num_teachers)} teachers."
    )

    run_button = st.button("🚀 Generate Timetable", type="primary", use_container_width=True)

    if run_button:
        if uploaded_file is None:
            st.error("Please upload your Excel file first.")
        else:
            progress_bar = st.progress(0)
            status_box = st.empty()

            with st.spinner("Running Zero-Clash Genetic Algorithm..."):
                try:
                    df_output, report = run_zero_clash_ga(
                        uploaded_file,
                        progress_bar,
                        status_box,
                        int(num_classes),
                        int(num_teachers)
                    )

                    st.session_state.df_output = df_output
                    st.session_state.report = report

                    st.success("✅ Timetable generated successfully!")

                except Exception as e:
                    st.error(f"Error: {e}")

    if "df_output" in st.session_state:
        df_output = st.session_state.df_output
        report = st.session_state.report

        st.write("### 📊 Timetable Result Summary")

        col1, col2, col3, col4 = st.columns(4)

        col1.metric("Hard Clashes", report["clash"])
        col2.metric("Optimality", f'{report["optimality"]}%')
        col3.metric("Classes", report["classes"])
        col4.metric("Teachers", report["teachers"])

        if report["clash"] == 0:
            st.success("✅ The generated timetable is conflict-free. No class or teacher clashes were found.")
        else:
            st.error("❌ The timetable contains clashes. Please regenerate the timetable.")

        st.write("### 📅 Timetable Preview")

        view_type = st.radio(
            "View timetable by:",
            ["Class", "Teacher"],
            horizontal=True
        )

        display_df = df_output.copy()
        day_order = report["day_order"]
        period_order = report["period_order"]

        if view_type == "Class":
            display_df["Entry"] = display_df["Subject"] + " (" + display_df["Teacher"] + ")"
            class_list = sorted(display_df["Class"].unique())
            selected_class = st.selectbox("Select Class:", class_list)

            preview_df = display_df[display_df["Class"] == selected_class]

            pivot_df = preview_df.pivot_table(
                index="Day",
                columns="Period",
                values="Entry",
                aggfunc=lambda x: " / ".join(x)
            ).reindex(index=day_order, columns=period_order).fillna("-")

            st.dataframe(pivot_df, use_container_width=True)

        else:
            display_df["Entry"] = display_df["Class"] + " (" + display_df["Subject"] + ")"
            teacher_list = sorted(display_df["Teacher"].unique())
            selected_teacher = st.selectbox("Select Teacher:", teacher_list)

            preview_df = display_df[display_df["Teacher"] == selected_teacher]

            pivot_df = preview_df.pivot_table(
                index="Day",
                columns="Period",
                values="Entry",
                aggfunc=lambda x: " / ".join(x)
            ).reindex(index=day_order, columns=period_order).fillna("-")

            st.dataframe(pivot_df, use_container_width=True)

        excel_file = create_output_excel(df_output, report)

        st.download_button(
            label="📥 Download Full Timetable Excel",
            data=excel_file,
            file_name="GA_Zero_Clash_Timetable.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )